from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
import unicodedata

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # Dependencia opcional para entornos sin paquetes instalados.
    Workbook = load_workbook = None
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
except ImportError:  # Dependencia opcional para entornos sin paquetes instalados.
    letter = None
    canvas = None

from .config import EXPORT_DIR
from .database import Database

INGREDIENT_SEARCH_ALIASES = {
    "pollo": ("pechuga", "muslo", "alita", "ala", "pierna", "contramuslo"),
    "res": ("carne", "lomo", "posta", "falda", "costilla"),
    "cerdo": ("tocino", "chuleta", "lomo", "costilla", "panceta"),
    "pescado": ("filete", "tilapia", "salmon", "atun", "trucha"),
}

def normalize_search(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text or "")
    return "".join(ch for ch in normalized if not unicodedata.combining(ch)).casefold().strip()

def ingredient_matches(name: str, search: str) -> bool:
    query = normalize_search(search)
    if not query:
        return True
    normalized_name = normalize_search(name)
    terms = [query, *INGREDIENT_SEARCH_ALIASES.get(query, ())]
    return any(term in normalized_name for term in terms)

@dataclass(frozen=True)
class PriceMetrics:
    total_cost: float
    cost_per_portion: float
    suggested_price: float
    sale_price: float
    profit_value: float
    profit_percent: float
    gross_margin: float
    profitability: float

def unit_cost(purchase_cost: float, purchased_quantity: float) -> float:
    if purchased_quantity <= 0:
        raise ValueError("La cantidad comprada debe ser mayor a cero")
    return round(purchase_cost / purchased_quantity, 6)

def price_metrics(total_cost: float, portions: float, margin: float, manual_price: float) -> PriceMetrics:
    cost_per_portion = total_cost / portions if portions else 0
    suggested = cost_per_portion * (1 + margin / 100)
    sale = manual_price if manual_price > 0 else suggested
    profit = sale - cost_per_portion
    profit_percent = (profit / cost_per_portion * 100) if cost_per_portion else 0
    gross_margin = (profit / sale * 100) if sale else 0
    profitability = (sale / cost_per_portion) if cost_per_portion else 0
    return PriceMetrics(total_cost, cost_per_portion, suggested, sale, profit, profit_percent, gross_margin, profitability)

class MahebeerService:
    def __init__(self, db: Database):
        self.db = db
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    def add_ingredient(self, name: str, cost: float, qty: float, measure: str) -> None:
        uc = unit_cost(cost, qty)
        with self.db.connect() as c:
            cur = c.execute("INSERT INTO ingredients(name,purchase_cost,purchased_quantity,measure_type,unit_cost) VALUES(?,?,?,?,?)", (name.strip(), cost, qty, measure, uc))
            self._history(c, "Ingrediente", cur.lastrowid, "Crear", name)

    def update_ingredient(self, iid: int, name: str, cost: float, qty: float, measure: str) -> None:
        uc = unit_cost(cost, qty)
        with self.db.connect() as c:
            c.execute("UPDATE ingredients SET name=?, purchase_cost=?, purchased_quantity=?, measure_type=?, unit_cost=?, updated_at=CURRENT_TIMESTAMP WHERE id=?", (name.strip(), cost, qty, measure, uc, iid))
            self._history(c, "Ingrediente", iid, "Editar", name)

    def soft_delete_ingredient(self, iid: int) -> None:
        with self.db.connect() as c:
            c.execute("UPDATE ingredients SET deleted_at=CURRENT_TIMESTAMP WHERE id=?", (iid,))
            self._history(c, "Ingrediente", iid, "Papelera", "Ingrediente eliminado")

    def restore(self, entity: str, eid: int) -> None:
        table = "ingredients" if entity == "Ingrediente" else "recipes"
        with self.db.connect() as c:
            c.execute(f"UPDATE {table} SET deleted_at=NULL, updated_at=CURRENT_TIMESTAMP WHERE id=?", (eid,))
            self._history(c, entity, eid, "Restaurar", "Restaurado desde papelera")

    def ingredients(self, search: str = "", include_deleted: bool = False):
        where = "WHERE 1=1" + ("" if include_deleted else " AND deleted_at IS NULL")
        with self.db.connect() as c:
            rows = c.execute(f"SELECT * FROM ingredients {where} ORDER BY name").fetchall()
        return [row for row in rows if ingredient_matches(row["name"], search)]

    def save_recipe(self, rid: int | None, name: str, description: str, portions: float, margin: float, manual_price: float, items: Iterable[dict]) -> int:
        with self.db.connect() as c:
            if rid:
                c.execute("UPDATE recipes SET name=?, description=?, yield_portions=?, suggested_margin=?, manual_sale_price=?, updated_at=CURRENT_TIMESTAMP WHERE id=?", (name.strip(), description, portions, margin, manual_price, rid))
                c.execute("DELETE FROM recipe_items WHERE recipe_id=?", (rid,))
                recipe_id = rid
                action = "Editar"
            else:
                cur = c.execute("INSERT INTO recipes(name,description,yield_portions,suggested_margin,manual_sale_price) VALUES(?,?,?,?,?)", (name.strip(), description, portions, margin, manual_price))
                recipe_id = cur.lastrowid
                action = "Crear"
            for item in items:
                ing = c.execute("SELECT unit_cost FROM ingredients WHERE id=?", (item["ingredient_id"],)).fetchone()
                c.execute("INSERT INTO recipe_items(recipe_id,ingredient_id,quantity,unit,unit_cost_snapshot) VALUES(?,?,?,?,?)", (recipe_id, item["ingredient_id"], item["quantity"], item["unit"], ing["unit_cost"]))
            self._history(c, "Receta", recipe_id, action, name)
            return recipe_id

    def recipes(self, search: str = "", include_deleted: bool = False):
        deleted = "" if include_deleted else " AND r.deleted_at IS NULL"
        with self.db.connect() as c:
            return c.execute(f"""
                SELECT r.*, COALESCE(SUM(ri.quantity*i.unit_cost),0) total_cost
                FROM recipes r LEFT JOIN recipe_items ri ON r.id=ri.recipe_id LEFT JOIN ingredients i ON i.id=ri.ingredient_id
                WHERE r.name LIKE ? {deleted} GROUP BY r.id ORDER BY r.name
            """, (f"%{search}%",)).fetchall()

    def recipe_items(self, rid: int):
        with self.db.connect() as c:
            return c.execute("""SELECT ri.*, i.name, i.unit_cost, ri.quantity*i.unit_cost cost FROM recipe_items ri JOIN ingredients i ON i.id=ri.ingredient_id WHERE recipe_id=? ORDER BY ri.id""", (rid,)).fetchall()

    def duplicate_recipe(self, rid: int) -> int:
        with self.db.connect() as c:
            r = c.execute("SELECT * FROM recipes WHERE id=?", (rid,)).fetchone()
            new_name = f"Copia de {r['name']}"
            cur = c.execute("INSERT INTO recipes(name,description,yield_portions,suggested_margin,manual_sale_price) VALUES(?,?,?,?,?)", (new_name, r['description'], r['yield_portions'], r['suggested_margin'], r['manual_sale_price']))
            new_id = cur.lastrowid
            c.execute("INSERT INTO recipe_items(recipe_id,ingredient_id,quantity,unit,unit_cost_snapshot) SELECT ?,ingredient_id,quantity,unit,unit_cost_snapshot FROM recipe_items WHERE recipe_id=?", (new_id, rid))
            self._history(c, "Receta", new_id, "Duplicar", new_name)
            return new_id

    def soft_delete_recipe(self, rid: int) -> None:
        with self.db.connect() as c:
            c.execute("UPDATE recipes SET deleted_at=CURRENT_TIMESTAMP WHERE id=?", (rid,))
            self._history(c, "Receta", rid, "Papelera", "Receta eliminada")

    def dashboard_reports(self):
        with self.db.connect() as c:
            return {
                "expensive": c.execute("SELECT name, unit_cost FROM ingredients WHERE deleted_at IS NULL ORDER BY unit_cost DESC LIMIT 10").fetchall(),
                "inventory": c.execute("SELECT COALESCE(SUM(purchase_cost),0) value FROM ingredients WHERE deleted_at IS NULL").fetchone()["value"],
                "avg_cost": c.execute("SELECT AVG(total) value FROM (SELECT SUM(ri.quantity*i.unit_cost) total FROM recipes r JOIN recipe_items ri ON r.id=ri.recipe_id JOIN ingredients i ON i.id=ri.ingredient_id WHERE r.deleted_at IS NULL GROUP BY r.id)").fetchone()["value"] or 0,
                "history": c.execute("SELECT * FROM change_history ORDER BY id DESC LIMIT 50").fetchall(),
            }

    def export_excel(self, path: Path) -> Path:
        if Workbook is None:
            path = path.with_suffix(".csv")
            rows = list(self.ingredients())
            with path.open("w", encoding="utf-8") as fh:
                if rows:
                    fh.write(",".join(rows[0].keys()) + "\n")
                    for r in rows:
                        fh.write(",".join(str(x) for x in r) + "\n")
            return path
        wb = Workbook()
        for title, rows in (("Ingredientes", self.ingredients()), ("Recetas", self.recipes())):
            ws = wb.active if title == "Ingredientes" else wb.create_sheet(title)
            ws.title = title
            if rows:
                ws.append(rows[0].keys())
                for r in rows: ws.append(list(r))
        wb.save(path)
        return path

    def import_ingredients_excel(self, path: Path) -> int:
        if load_workbook is None:
            raise RuntimeError("Instale openpyxl para importar archivos Excel")
        wb = load_workbook(path)
        ws = wb.active
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            self.add_ingredient(str(row[0]), float(row[1]), float(row[2]), str(row[3]))
            count += 1
        return count

    def export_pdf(self, path: Path) -> Path:
        if canvas is None:
            path = path.with_suffix(".txt")
            path.write_text("\n".join(r["name"] for r in self.recipes()), encoding="utf-8")
            return path
        pdf = canvas.Canvas(str(path), pagesize=letter)
        y = 750
        pdf.setTitle("Reporte MAHEBEER")
        pdf.drawString(40, y, "MAHEBEER - Reporte de Recetas y Costos")
        y -= 30
        for r in self.recipes():
            m = price_metrics(r["total_cost"], r["yield_portions"], r["suggested_margin"], r["manual_sale_price"])
            pdf.drawString(40, y, f"{r['name']} | Costo: ${m.total_cost:,.0f} | Sugerido: ${m.suggested_price:,.0f} | Margen: {m.gross_margin:.1f}%")
            y -= 18
            if y < 50:
                pdf.showPage(); y = 750
        pdf.save()
        return path

    def _history(self, c, entity: str, eid: int, action: str, details: str) -> None:
        c.execute("INSERT INTO change_history(entity,entity_id,action,details) VALUES(?,?,?,?)", (entity, eid, action, details))
