from __future__ import annotations

import sys
from pathlib import Path
from datetime import datetime

from PySide6.QtWidgets import (
    QApplication, QComboBox, QDialog, QFileDialog, QFormLayout, QHBoxLayout,
    QLabel, QLineEdit, QMainWindow, QMessageBox, QPushButton, QDoubleSpinBox,
    QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget, QCheckBox
)

from .app import MainWindow as BaseMainWindow, STYLE, money
from .config import EXPORT_DIR


PRODUCT_SCHEMA = """
CREATE TABLE IF NOT EXISTS departments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL COLLATE NOCASE UNIQUE,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS products (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    barcode TEXT COLLATE NOCASE UNIQUE,
    name TEXT NOT NULL COLLATE NOCASE UNIQUE,
    sale_type TEXT NOT NULL DEFAULT 'Por Unidad/Pza',
    cost_price REAL NOT NULL DEFAULT 0,
    sale_price REAL NOT NULL DEFAULT 0,
    wholesale_price REAL NOT NULL DEFAULT 0,
    department_id INTEGER REFERENCES departments(id) ON DELETE SET NULL,
    uses_inventory INTEGER NOT NULL DEFAULT 0,
    current_stock REAL NOT NULL DEFAULT 0,
    minimum_stock REAL NOT NULL DEFAULT 0,
    deleted_at TEXT,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS idx_products_name ON products(name);
CREATE INDEX IF NOT EXISTS idx_products_barcode ON products(barcode);
"""


class ProductDialog(QDialog):
    def __init__(self, db, parent=None, row=None):
        super().__init__(parent)
        self.db = db
        self.row = row
        self.setWindowTitle("Nuevo producto" if row is None else "Modificar producto")
        self.resize(650, 520)

        self.barcode = QLineEdit(row["barcode"] or "" if row else "")
        self.name = QLineEdit(row["name"] if row else "")
        self.sale_type = QComboBox()
        self.sale_type.addItems(["Por Unidad/Pza", "A Granel (Usa Decimales)", "Como paquete (kit)"])
        if row:
            self.sale_type.setCurrentText(row["sale_type"])

        self.cost = QDoubleSpinBox(); self._money_spin(self.cost, row["cost_price"] if row else 0)
        self.sale = QDoubleSpinBox(); self._money_spin(self.sale, row["sale_price"] if row else 0)
        self.wholesale = QDoubleSpinBox(); self._money_spin(self.wholesale, row["wholesale_price"] if row else 0)

        self.department = QComboBox()
        self._load_departments(row["department_id"] if row else None)

        self.uses_inventory = QCheckBox("Este producto SI utiliza inventario.")
        self.uses_inventory.setChecked(bool(row["uses_inventory"]) if row else False)
        self.current_stock = QDoubleSpinBox(); self._qty_spin(self.current_stock, row["current_stock"] if row else 0)
        self.minimum_stock = QDoubleSpinBox(); self._qty_spin(self.minimum_stock, row["minimum_stock"] if row else 0)

        form = QFormLayout()
        form.addRow("Código de barras", self.barcode)
        form.addRow("Descripción", self.name)
        form.addRow("Se vende", self.sale_type)
        form.addRow("Precio Costo", self.cost)
        form.addRow("Precio Venta", self.sale)
        form.addRow("Precio Mayoreo", self.wholesale)
        form.addRow("Departamento", self.department)
        form.addRow("Inventario", self.uses_inventory)
        form.addRow("Cantidad Actual", self.current_stock)
        form.addRow("Mínimo", self.minimum_stock)

        save = QPushButton("✓ Guardar Producto")
        cancel = QPushButton("✕ Cancelar")
        save.clicked.connect(self._save)
        cancel.clicked.connect(self.reject)
        buttons = QHBoxLayout(); buttons.addWidget(save); buttons.addStretch(); buttons.addWidget(cancel)

        layout = QVBoxLayout(self)
        title = QLabel("NUEVO PRODUCTO" if row is None else "MODIFICAR PRODUCTO")
        title.setStyleSheet("font-size:22px;font-weight:800;color:#fbbf24;padding:4px 0;")
        layout.addWidget(title)
        layout.addLayout(form)
        layout.addStretch()
        layout.addLayout(buttons)

    @staticmethod
    def _money_spin(w, value):
        w.setMaximum(999999999); w.setDecimals(0); w.setPrefix("$"); w.setValue(float(value or 0))

    @staticmethod
    def _qty_spin(w, value):
        w.setMaximum(999999999); w.setDecimals(2); w.setValue(float(value or 0))

    def _load_departments(self, selected):
        self.department.clear()
        self.department.addItem("- Sin Departamento -", None)
        with self.db.connect() as c:
            rows = c.execute("SELECT id,name FROM departments ORDER BY name").fetchall()
        for r in rows:
            self.department.addItem(r["name"], r["id"])
        if selected is not None:
            idx = self.department.findData(selected)
            if idx >= 0: self.department.setCurrentIndex(idx)

    def _save(self):
        if not self.name.text().strip():
            QMessageBox.warning(self, "Producto", "Escriba la descripción del producto.")
            self.name.setFocus(); return
        if self.sale.value() < 0 or self.cost.value() < 0:
            QMessageBox.warning(self, "Producto", "Los precios no pueden ser negativos.")
            return
        self.accept()


class DepartmentDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db; self.setWindowTitle("Departamentos"); self.resize(430, 360)
        self.table = QTableWidget(0, 2); self.table.setHorizontalHeaderLabels(["ID", "Departamento"])
        self.name = QLineEdit(); self.name.setPlaceholderText("Nombre del departamento")
        add = QPushButton("Agregar"); delete = QPushButton("Eliminar")
        add.clicked.connect(self.add); delete.clicked.connect(self.delete)
        row = QHBoxLayout(); row.addWidget(self.name); row.addWidget(add); row.addWidget(delete)
        lay = QVBoxLayout(self); lay.addLayout(row); lay.addWidget(self.table); self.load()

    def load(self):
        self.table.setRowCount(0)
        with self.db.connect() as c: rows = c.execute("SELECT id,name FROM departments ORDER BY name").fetchall()
        for r in rows:
            i = self.table.rowCount(); self.table.insertRow(i); self.table.setItem(i,0,QTableWidgetItem(str(r["id"]))); self.table.setItem(i,1,QTableWidgetItem(r["name"]))

    def add(self):
        name = self.name.text().strip()
        if not name: return
        try:
            with self.db.connect() as c: c.execute("INSERT INTO departments(name) VALUES(?)", (name,))
        except Exception as e: QMessageBox.warning(self,"Departamento",f"No se pudo guardar: {e}")
        self.name.clear(); self.load()

    def delete(self):
        row = self.table.currentRow()
        if row < 0: return
        did = int(self.table.item(row,0).text())
        with self.db.connect() as c: c.execute("DELETE FROM departments WHERE id=?", (did,))
        self.load()


class ProductsPage(QWidget):
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        with self.db.connect() as c: c.executescript(PRODUCT_SCHEMA)
        self._build(); self.load()

    def _build(self):
        title = QLabel("PRODUCTOS")
        title.setStyleSheet("font-size:24px;font-weight:800;color:#fbbf24;padding:5px 0;")
        self.search = QLineEdit(); self.search.setPlaceholderText("Buscar producto por código de barras o descripción...")
        self.search.textChanged.connect(self.load)

        buttons = QHBoxLayout()
        for text, fn in [
            ("Nuevo", self.new), ("Modificar", self.edit), ("Eliminar", self.delete),
            ("Departamentos", self.departments), ("Ventas por Periodo", self.sales_period),
            ("Promociones", self.promotions), ("Importar ...", self.import_products)
        ]:
            b = QPushButton(text); b.clicked.connect(fn); buttons.addWidget(b)
        buttons.addStretch()

        self.table = QTableWidget(0, 9)
        self.table.setHorizontalHeaderLabels([
            "ID", "Código de barras", "Descripción", "Precio Costo", "Precio Venta",
            "Mayoreo", "Departamento", "Inventario", "Existencia"
        ])
        self.table.setAlternatingRowColors(True); self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.doubleClicked.connect(lambda _: self.edit())

        layout = QVBoxLayout(self); layout.addWidget(title); layout.addLayout(buttons); layout.addWidget(self.search); layout.addWidget(self.table)

    def load(self):
        term = f"%{self.search.text().strip()}%"
        with self.db.connect() as c:
            rows = c.execute("""
                SELECT p.*, COALESCE(d.name,'- Sin Departamento -') department
                FROM products p LEFT JOIN departments d ON d.id=p.department_id
                WHERE p.deleted_at IS NULL AND (p.name LIKE ? OR COALESCE(p.barcode,'') LIKE ?)
                ORDER BY p.name
            """, (term,term)).fetchall()
        self.table.setRowCount(0)
        for r in rows:
            i=self.table.rowCount(); self.table.insertRow(i)
            vals=[r["id"],r["barcode"] or "",r["name"],money(r["cost_price"]),money(r["sale_price"]),money(r["wholesale_price"]),r["department"],"Sí" if r["uses_inventory"] else "No",f'{r["current_stock"]:.2f}']
            for col,val in enumerate(vals): self.table.setItem(i,col,QTableWidgetItem(str(val)))

    def selected_id(self):
        row=self.table.currentRow()
        return int(self.table.item(row,0).text()) if row>=0 and self.table.item(row,0) else None

    def get_row(self, pid):
        with self.db.connect() as c: return c.execute("SELECT * FROM products WHERE id=?",(pid,)).fetchone()

    def new(self):
        d=ProductDialog(self.db,self)
        if d.exec():
            self.save_dialog(d,None); self.load()

    def edit(self):
        pid=self.selected_id()
        if not pid: QMessageBox.information(self,"Productos","Seleccione un producto."); return
        d=ProductDialog(self.db,self,self.get_row(pid))
        if d.exec(): self.save_dialog(d,pid); self.load()

    def save_dialog(self,d,pid):
        values=(d.barcode.text().strip() or None,d.name.text().strip(),d.sale_type.currentText(),d.cost.value(),d.sale.value(),d.wholesale.value(),d.department.currentData(),int(d.uses_inventory.isChecked()),d.current_stock.value(),d.minimum_stock.value())
        try:
            with self.db.connect() as c:
                if pid:
                    c.execute("""UPDATE products SET barcode=?,name=?,sale_type=?,cost_price=?,sale_price=?,wholesale_price=?,department_id=?,uses_inventory=?,current_stock=?,minimum_stock=?,updated_at=CURRENT_TIMESTAMP WHERE id=?""", values+(pid,))
                else:
                    c.execute("""INSERT INTO products(barcode,name,sale_type,cost_price,sale_price,wholesale_price,department_id,uses_inventory,current_stock,minimum_stock) VALUES(?,?,?,?,?,?,?,?,?,?)""", values)
        except Exception as e: QMessageBox.warning(self,"Producto",f"No se pudo guardar el producto.\n\n{e}")

    def delete(self):
        pid=self.selected_id()
        if not pid: return
        if QMessageBox.question(self,"Eliminar producto","¿Desea enviar este producto a la papelera?") != QMessageBox.Yes: return
        with self.db.connect() as c: c.execute("UPDATE products SET deleted_at=? WHERE id=?",(datetime.now().isoformat(timespec="seconds"),pid))
        self.load()

    def departments(self):
        d=DepartmentDialog(self.db,self); d.exec(); self.load()

    def sales_period(self):
        QMessageBox.information(self,"Ventas por Periodo","Este botón queda preparado para enlazar los productos con el módulo Ventas.")

    def promotions(self):
        QMessageBox.information(self,"Promociones","Este botón queda preparado para el módulo de promociones.")

    def import_products(self):
        path,_=QFileDialog.getOpenFileName(self,"Importar productos",str(Path.home()),"Excel (*.xlsx);;CSV (*.csv)")
        if not path: return
        if path.lower().endswith('.xlsx'):
            try:
                from openpyxl import load_workbook
                ws=load_workbook(path,read_only=True,data_only=True).active
                rows=list(ws.iter_rows(values_only=True))
                if not rows: return
                headers=[str(x or '').strip().lower() for x in rows[0]]
                idx={h:i for i,h in enumerate(headers)}
                imported=0
                with self.db.connect() as c:
                    for r in rows[1:]:
                        name=r[idx.get('descripción',idx.get('descripcion',0))] if r else None
                        if not name: continue
                        barcode=r[idx.get('código de barras',idx.get('codigo de barras',1))] if 'código de barras' in idx or 'codigo de barras' in idx else None
                        cost=r[idx.get('precio costo',3)] if 'precio costo' in idx else 0
                        sale=r[idx.get('precio venta',4)] if 'precio venta' in idx else 0
                        c.execute("INSERT OR IGNORE INTO products(barcode,name,cost_price,sale_price) VALUES(?,?,?,?)",(str(barcode) if barcode else None,str(name),float(cost or 0),float(sale or 0))); imported+=1
                self.load(); QMessageBox.information(self,"Importar",f"Productos importados: {imported}")
            except Exception as e: QMessageBox.warning(self,"Importar",f"No se pudo importar:\n{e}")


class MainWindow(BaseMainWindow):
    def __init__(self):
        super().__init__()
        self.products_page=ProductsPage(self.db,self)
        self.centralWidget().insertTab(0,self.products_page,"Productos")
        self.centralWidget().setCurrentIndex(0)
        self.setWindowTitle("MAHEBEER - Productos")


def main():
    app=QApplication(sys.argv); app.setStyleSheet(STYLE); w=MainWindow(); w.show(); sys.exit(app.exec())


if __name__ == "__main__": main()
